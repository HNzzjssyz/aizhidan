import json
import os
import random
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

DATA_DIR = os.path.dirname(os.path.abspath(__file__))

with open(os.path.join(DATA_DIR, "product_pinyin_2026-05-16.json"), "r", encoding="utf-8") as f:
    product_data = json.load(f)
with open(os.path.join(DATA_DIR, "customer_pinyin_2026-05-16.json"), "r", encoding="utf-8") as f:
    customer_data = json.load(f)

products = product_data["products"]
customers = customer_data["customers"]

random.seed(42)

CONTACTS = ["李总", "陈经理", "张经理", "赵主任", "王工", "刘总", "周经理", "吴主管"]
PHONES = [
    "13508157429", "17853573823", "18507437181", "13964411347",
    "18524806516", "18726563708", "18197613238", "13513140753",
    "19906248900", "13875340444", "13699251800", "15201239876",
]
ADDRESSES = [
    "深圳市南山区科技园路66号", "北京市海淀区中关村大街1号",
    "广州市天河区天河路385号", "上海市浦东新区张江路88号",
    "杭州市西湖区文三路200号", "南京市鼓楼区汉中路100号",
]
REMARKS = ["", "月结", "加急", "需开票", "常规", "货到付款"]
UNITS = ["个", "台", "件", "套", "箱"]

VERBAL_TEMPLATES = [
    lambda c, items: f"帮我给{c}发" + "、".join(f"{it['qty']}{it['unit']}{it['name']}{' '+it['model'] if it['model'] else ''}" for it in items),
    lambda c, items: "、".join(f"{it['name']}{' '+it['model'] if it['model'] else ''} {it['qty']}{it['unit']}" for it in items) + f"，给{c}",
    lambda c, items: f"客户{c}要" + "，".join(f"{it['qty']}{it['unit']}{it['name']}" for it in items),
    lambda c, items: f"下个单：{c}，" + "，".join(f"{it['name']}{' '+it['model'] if it['model'] else ''} {it['qty']}{it['unit']}" for it in items),
    lambda c, items: c + "\n" + "\n".join(f"{it['name']}\n{it['model']}\n{it['qty']}{it['unit']}" if it['model'] else f"{it['name']}\n{it['qty']}{it['unit']}" for it in items),
    lambda c, items: "对，" + "，".join(f"{it['name']}{' '+it['model'] if it['model'] else ''}，{it['qty']}{it['unit']}" for it in items) + f"，客户{c}",
    lambda c, items: f"给{c}出个单，" + "，".join(f"{it['model']} {it['name']} {it['qty']}{it['unit']}" if it['model'] else f"{it['name']} {it['qty']}{it['unit']}" for it in items),
    lambda c, items: "订单：" + "|".join(f"{it['name']}|{it['model']}|{it['qty']}" if it['model'] else f"{it['name']}||{it['qty']}" for it in items) + f"|{c}",
]

selected_customers = random.sample([c for c in customers if len(c["name"]) >= 4 and "测试" not in c["name"] and "演示" not in c["name"]], 15)

selected_products = random.sample(products, min(40, len(products)))

verbal_orders = []
screenshot_orders = []

order_date = datetime.date(2026, 5, 18)

for i in range(15):
    customer = selected_customers[i]
    num_products = random.randint(2, 5)
    order_products = random.sample(selected_products, num_products)

    items = []
    for p in order_products:
        qty = random.randint(1, 30)
        unit = random.choice(UNITS)
        price = round(random.uniform(500, 50000), 2)
        items.append({
            "product_id": p["id"],
            "name": p["name"],
            "model": p["model"],
            "spec": p["spec"],
            "qty": qty,
            "unit": unit,
            "price": price,
            "amount": round(qty * price, 2),
        })

    date_str = (order_date + datetime.timedelta(days=i)).strftime("%Y-%m-%d")
    order_no = f"ORD{(order_date + datetime.timedelta(days=i)).strftime('%Y%m%d')}{i+1:03d}"
    contact = random.choice(CONTACTS)
    phone = random.choice(PHONES)
    address = random.choice(ADDRESSES)
    remark = random.choice(REMARKS)

    template = random.choice(VERBAL_TEMPLATES)
    content = template(customer["name"], items)

    verbal_type = random.choice(["微信聊天", "电话口述", "短信/钉钉", "语音转文字(含错字)"])

    verbal_orders.append({
        "id": f"VO-{i+1:03d}",
        "type": verbal_type,
        "date": date_str,
        "content": content,
        "expected_customer": customer["name"],
        "expected_customer_id": customer["id"],
        "expected_products": [
            {
                "product_id": it["product_id"],
                "name": it["name"],
                "model": it["model"],
                "spec": it["spec"],
                "qty": it["qty"],
                "unit": it["unit"],
            }
            for it in items
        ],
        "note": f"多产品订单：{len(items)}个产品",
    })

    screenshot_orders.append({
        "id": f"SS-{i+1:03d}",
        "date": date_str,
        "order_no": order_no,
        "customer_name": customer["name"],
        "customer_id": customer["id"],
        "contact": contact,
        "phone": phone,
        "address": address,
        "remark": remark,
        "products": [
            {
                "product_id": it["product_id"],
                "name": it["name"],
                "model": it["model"],
                "spec": it["spec"],
                "qty": it["qty"],
                "unit": it["unit"],
                "price": it["price"],
                "amount": it["amount"],
            }
            for it in items
        ],
        "total_amount": round(sum(it["amount"] for it in items), 2),
    })

verbal_output = {
    "description": "模拟口语文字订单数据 - 用于拼音匹配实验测试（多产品订单）",
    "generatedAt": datetime.datetime.now().isoformat(),
    "totalOrders": len(verbal_orders),
    "data": verbal_orders,
}
with open(os.path.join(DATA_DIR, "verbal_orders_test.json"), "w", encoding="utf-8") as f:
    json.dump(verbal_output, f, ensure_ascii=False, indent=2)
print(f"verbal_orders_test.json: {len(verbal_orders)} orders generated")

screenshot_output = {
    "description": "模拟截图订单数据 - 用于拼音匹配实验测试（多产品订单）",
    "generatedAt": datetime.datetime.now().isoformat(),
    "totalOrders": len(screenshot_orders),
    "data": screenshot_orders,
}
with open(os.path.join(DATA_DIR, "screenshot_orders_test.json"), "w", encoding="utf-8") as f:
    json.dump(screenshot_output, f, ensure_ascii=False, indent=2)
print(f"screenshot_orders_test.json: {len(screenshot_orders)} orders generated")

BADGE_MAP = {
    "月结": "badge-monthly",
    "加急": "badge-urgent",
    "需开票": "badge-invoice",
    "常规": "badge-normal",
    "货到付款": "badge-monthly",
}

html_parts = [
    '<!DOCTYPE html>',
    '<html lang="zh-CN">',
    '<head>',
    '<meta charset="UTF-8">',
    '<title>模拟订单截图数据</title>',
    '<style>',
    'body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif; background: #f0f2f5; padding: 20px; }',
    '.screenshot-container { max-width: 800px; margin: 0 auto; }',
    '.order-card { background: #fff; border-radius: 8px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); margin-bottom: 24px; overflow: hidden; }',
    '.order-header { background: #1890ff; color: #fff; padding: 12px 20px; font-size: 16px; font-weight: 600; display: flex; justify-content: space-between; }',
    '.order-header .order-no { font-size: 13px; opacity: 0.9; }',
    '.order-body { padding: 16px 20px; }',
    '.info-row { display: flex; margin-bottom: 8px; font-size: 14px; line-height: 1.8; }',
    '.info-label { color: #999; width: 80px; flex-shrink: 0; }',
    '.info-value { color: #333; font-weight: 500; }',
    '.product-table { width: 100%; border-collapse: collapse; margin-top: 12px; font-size: 13px; }',
    '.product-table th { background: #fafafa; padding: 8px 12px; text-align: left; border: 1px solid #e8e8e8; color: #666; font-weight: 500; }',
    '.product-table td { padding: 8px 12px; border: 1px solid #e8e8e8; color: #333; }',
    '.product-table .amount { text-align: right; font-weight: 600; color: #f5222d; }',
    '.order-footer { padding: 12px 20px; background: #fafafa; border-top: 1px solid #e8e8e8; display: flex; justify-content: space-between; font-size: 13px; color: #666; }',
    '.badge { display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 12px; margin-left: 8px; }',
    '.badge-urgent { background: #fff1f0; color: #f5222d; border: 1px solid #ffa39e; }',
    '.badge-normal { background: #f6ffed; color: #52c41a; border: 1px solid #b7eb8f; }',
    '.badge-invoice { background: #e6f7ff; color: #1890ff; border: 1px solid #91d5ff; }',
    '.badge-monthly { background: #fff7e6; color: #fa8c16; border: 1px solid #ffd591; }',
    'h2 { color: #333; margin-bottom: 16px; }',
    '</style>',
    '</head>',
    '<body>',
    '<div class="screenshot-container">',
    f'<h2>📋 模拟订单截图数据（共{len(screenshot_orders)}条）</h2>',
]

for order in screenshot_orders:
    badge_html = ""
    if order["remark"] and order["remark"] in BADGE_MAP:
        badge_html = f'<span class="badge {BADGE_MAP[order["remark"]]}">{order["remark"]}</span>'

    html_parts.append('<div class="order-card">')
    html_parts.append(f'  <div class="order-header">')
    html_parts.append(f'    <span>销售订单{badge_html}</span>')
    html_parts.append(f'    <span class="order-no">{order["order_no"]}</span>')
    html_parts.append(f'  </div>')
    html_parts.append(f'  <div class="order-body">')
    html_parts.append(f'    <div class="info-row"><span class="info-label">客户名称</span><span class="info-value">{order["customer_name"]}</span></div>')
    html_parts.append(f'    <div class="info-row"><span class="info-label">联系人</span><span class="info-value">{order["contact"]} {order["phone"]}</span></div>')
    html_parts.append(f'    <div class="info-row"><span class="info-label">收货地址</span><span class="info-value">{order["address"]}</span></div>')
    html_parts.append(f'    <table class="product-table">')
    html_parts.append(f'      <tr><th>产品名称</th><th>型号</th><th>规格</th><th>数量</th><th>单价</th><th>金额</th></tr>')

    for p in order["products"]:
        html_parts.append(f'      <tr>')
        html_parts.append(f'        <td>{p["name"]}</td>')
        html_parts.append(f'        <td>{p["model"]}</td>')
        html_parts.append(f'        <td>{p["spec"]}</td>')
        html_parts.append(f'        <td>{p["qty"]}{p["unit"]}</td>')
        html_parts.append(f'        <td>¥{p["price"]:,.2f}</td>')
        html_parts.append(f'        <td class="amount">¥{p["amount"]:,.2f}</td>')
        html_parts.append(f'      </tr>')

    html_parts.append(f'    </table>')
    html_parts.append(f'  </div>')
    html_parts.append(f'  <div class="order-footer">')
    html_parts.append(f'    <span>下单日期：{order["date"]}</span>')
    html_parts.append(f'    <span>合计：¥{order["total_amount"]:,.2f}</span>')
    html_parts.append(f'  </div>')
    html_parts.append(f'</div>')

html_parts.append('</div>')
html_parts.append('</body>')
html_parts.append('</html>')

with open(os.path.join(DATA_DIR, "screenshot_orders_test.html"), "w", encoding="utf-8") as f:
    f.write("\n".join(html_parts))
print(f"screenshot_orders_test.html: {len(screenshot_orders)} orders generated")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)
header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
title_font = Font(bold=True, size=14)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')
right_align = Alignment(horizontal='right', vertical='center')

excel_dir = os.path.join(DATA_DIR, "excel_orders")
os.makedirs(excel_dir, exist_ok=True)

for order in screenshot_orders:
    wb = Workbook()
    ws = wb.active
    ws.title = "销售订单"

    ws.merge_cells('A1:F1')
    ws['A1'] = '销售订单'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align

    ws.merge_cells('A2:F2')
    ws['A2'] = f'订单编号：{order["order_no"]}'
    ws['A2'].alignment = left_align

    info_rows = [
        ('客户名称：', order["customer_name"]),
        ('联系人：', f'{order["contact"]} {order["phone"]}'),
        ('收货地址：', order["address"]),
        ('下单日期：', order["date"]),
    ]
    for idx, (label, value) in enumerate(info_rows, start=3):
        ws.merge_cells(f'A{idx}:B{idx}')
        ws[f'A{idx}'] = label
        ws[f'A{idx}'].font = Font(bold=True, size=11)
        ws[f'A{idx}'].alignment = left_align
        ws.merge_cells(f'C{idx}:F{idx}')
        ws[f'C{idx}'] = value
        ws[f'C{idx}'].alignment = left_align

    product_header_row = 7
    headers = ['产品名称', '型号', '规格', '数量', '单价', '金额']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=product_header_row, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    for p_idx, p in enumerate(order["products"]):
        row = product_header_row + 1 + p_idx
        values = [p["name"], p["model"], p["spec"], p["qty"], p["price"], p["amount"]]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = thin_border
            if col_idx in (4,):
                cell.alignment = center_align
            elif col_idx in (5, 6):
                cell.number_format = '#,##0.00'
                cell.alignment = right_align
            else:
                cell.alignment = left_align

    total_row = product_header_row + 1 + len(order["products"])
    ws.merge_cells(f'A{total_row}:D{total_row}')
    ws[f'A{total_row}'] = '合计'
    ws[f'A{total_row}'].font = Font(bold=True, size=11)
    ws[f'A{total_row}'].alignment = center_align
    ws[f'A{total_row}'].border = thin_border
    for col in range(2, 5):
        ws.cell(row=total_row, column=col).border = thin_border

    ws.cell(row=total_row, column=5).border = thin_border
    total_amount_cell = ws.cell(row=total_row, column=6, value=order["total_amount"])
    total_amount_cell.font = Font(bold=True, size=11, color="FF0000")
    total_amount_cell.number_format = '#,##0.00'
    total_amount_cell.alignment = right_align
    total_amount_cell.border = thin_border

    if order["remark"]:
        remark_row = total_row + 1
        ws.merge_cells(f'A{remark_row}:F{remark_row}')
        ws[f'A{remark_row}'] = f'备注：{order["remark"]}'
        ws[f'A{remark_row}'].alignment = left_align

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16

    filename = f"{order['order_no']}.xlsx"
    filepath = os.path.join(excel_dir, filename)
    wb.save(filepath)

print(f"excel_orders/: {len(screenshot_orders)} xlsx files generated in {excel_dir}")
print("All test data generated successfully!")
